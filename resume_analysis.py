import os
import time
import json
import hashlib
import threading
import logging
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from langchain_openai import ChatOpenAI, OpenAIEmbeddings
from langchain_core.output_parsers import JsonOutputParser
from langchain_core.prompts import ChatPromptTemplate
import fitz  # PyMuPDF
import csv
from openpyxl import load_workbook, Workbook
from langchain_community.vectorstores import FAISS
from langchain_community.docstore.in_memory import InMemoryDocstore
from langchain.schema import Document
from dotenv import load_dotenv
import asyncio
import faiss

# 환경 변수 로드
load_dotenv()

# 로그 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 환경 변수 사용
openai_api_key = os.getenv('OPENAI_API_KEY')
directory_to_watch = os.getenv('DIRECTORY_TO_WATCH')
excel_file_path = os.getenv('EXCEL_FILE_PATH')
hash_record_path = os.getenv('HASH_RECORD_PATH')
csv_file_path = os.getenv('CSV_FILE_PATH')
faiss_index_path = os.getenv('FAISS_INDEX_PATH')

# 특정 디렉토리 설정
DIRECTORY_TO_WATCH = directory_to_watch
EXCEL_FILE_PATH = excel_file_path
HASH_RECORD_PATH = hash_record_path
CSV_FILE_PATH = csv_file_path
FAISS_INDEX_PATH = faiss_index_path

# 해시 기록을 저장하기 위한 파일 로드
if os.path.exists(HASH_RECORD_PATH):
    with open(HASH_RECORD_PATH, "r") as f:
        processed_files = json.load(f)
else:
    processed_files = {}

# 큐 설정
file_queue = asyncio.Queue()

# FAISS 벡터 저장소 초기화
embeddings = OpenAIEmbeddings(openai_api_key=openai_api_key)
dimension = len(embeddings.embed_query("test"))
index = faiss.IndexFlatL2(dimension)

if os.path.exists(FAISS_INDEX_PATH):
    try:
        vector_store = FAISS.load_local(FAISS_INDEX_PATH, embeddings, allow_dangerous_deserialization=True)
        logging.info("Loaded FAISS index from disk.")
    except Exception as e:
        logging.error(f"Failed to load FAISS index: {e}")
        vector_store = FAISS(embedding_function=embeddings, index=index, docstore=InMemoryDocstore({}), index_to_docstore_id={})
else:
    vector_store = FAISS(embedding_function=embeddings, index=index, docstore=InMemoryDocstore({}), index_to_docstore_id={})

class ResumeHandler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory:
            return None
        elif event.src_path.endswith(".pdf"):
            asyncio.run_coroutine_threadsafe(file_queue.put(event.src_path), loop)

async def process_files_from_queue():
    while True:
        file_path = await file_queue.get()
        if file_path is None:
            break
        await process_resume(file_path)
        file_queue.task_done()

async def process_resume(file_path):
    file_hash = calculate_file_hash(file_path)

    if file_hash in processed_files:
        logging.info(f"File {file_path} has already been processed.")
        return

    file_info = parse_filename(file_path)
    if file_info:
        combined_text = extract_text_from_combined_files(file_info['applicant_name'])
        analysis_result = analyze_resume_with_langchain(combined_text, file_info)
        if analysis_result:
            update_excel_with_result(analysis_result)
            await update_faiss_with_result(analysis_result)
            processed_files[file_hash] = file_path
            save_hash_records()
        else:
            logging.error(f"Error processing file {file_path}.")
    else:
        logging.error(f"Error parsing filename {file_path}.")

def calculate_file_hash(file_path):
    hasher = hashlib.sha256()
    with open(file_path, "rb") as f:
        buf = f.read()
        hasher.update(buf)
    return hasher.hexdigest()

def save_hash_records():
    with open(HASH_RECORD_PATH, "w") as f:
        json.dump(processed_files, f)

def parse_filename(file_path):
    filename = os.path.basename(file_path)
    parts = filename.split('_')

    document_type = parts[0].split(')')[0].strip('(')
    applicant_name = parts[0].split(')')[1].strip('(')
    application_route = parts[1]
    date = parts[2]
    team_info = parts[3].split()
    team_name = team_info[0]
    if len(team_info) > 1:
        part_name = team_info[-2]
        position = team_info[-1]
        print("hhhhhh: ", position)
    else:
        part_name = ""
        position = ""

    return {
        'document_type': document_type,
        'applicant_name': applicant_name,
        'application_route': application_route,
        'date': date,
        'team_name': team_name,
        'part_name': part_name,
        'position': position
    }

def extract_text_from_combined_files(applicant_name):
    directory = os.listdir(DIRECTORY_TO_WATCH)
    combined_text = ""
    for file in directory:
        if applicant_name in file and file.endswith(".pdf"):
            file_path = os.path.join(DIRECTORY_TO_WATCH, file)
            combined_text += extract_text_from_pdf(file_path) + "\n\n"
    return combined_text

def analyze_resume_with_langchain(resume_text, file_info):
    prompt = ChatPromptTemplate.from_template(
        """
        이력서 내용: {RESUME_TEXT}
        이력서에서 다음 정보를 추출하여 요약해줘. 요약할 때 지원자의 특징이 잘 표현될 수 있도록 풍부하게 요약해줘. 그리고 마지막으로 JSON 형식으로 반환하세요.
        - 지원자 이름
        - 나이
        - 경력
        - 핵심기술력
        - 특징
        - 지원팀명
        - 지원파트
        - 직급

        JSON 형식 예시:
        {{
            "지원자 이름": "홍길동",
            "나이": "30",
            "지원팀명": "하이브리드 플랫폼팀",
            "지원파트": "백앤드 개발",
            "직급": "수석",
            "경력": "10년",
            "핵심기술력": "Python, Machine Learning",
            "특징": "개발능력이 우수하고 팀 리더 경험이 풍부하다."
        }}
        """
    )

    chain = (
            prompt
            | ChatOpenAI(
        temperature=0,
        model="gpt-3.5-turbo"
    )
            | JsonOutputParser()
    )

    result = chain.invoke({
        "RESUME_TEXT": resume_text + '"지원자 이름":' + file_info['applicant_name']
                       + ', "지원팀명":' + file_info['team_name']
                       + ', "지원파트":' + file_info['part_name']
                       + ', "직급":' + file_info['position']
    })
    if isinstance(result, dict):
        result.update(file_info)  # 파일 정보를 결과에 추가
        logging.info(f"Processed resume: {result}")
        return result
    else:
        logging.error("Error: Unexpected output format.")
        return {}

def extract_text_from_pdf(file_path):
    # PyMuPDF를 사용하여 PDF에서 텍스트 추출
    doc = fitz.open(file_path)
    text = ""
    for page in doc:
        text += page.get_text()
    return text

def update_excel_with_result(result):
    lock = threading.Lock()
    with lock:
        # 엑셀 파일 업데이트
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = load_workbook(EXCEL_FILE_PATH)
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["지원자 이름", "나이", "지원팀명", "지원파트", "직급", "경력", "핵심기술력", "특징"])  # 헤더 추가

        sheet = workbook.active
        # JSON 데이터를 문자열로 변환하여 저장
        name = result.get("지원자 이름", "")
        age = result.get("나이", "")
        team_name = result.get("지원팀명", "")
        part_name = result.get("지원파트", "")
        position = result.get("직급", "")
        experience = result.get("경력", "")
        skills = result.get("핵심기술력", "")
        characteristics = result.get("특징", "")

        sheet.append([name, age, team_name, part_name, position, experience, skills, characteristics])
        workbook.save(EXCEL_FILE_PATH)

def update_csv_with_result(result):
    # CSV 파일 업데이트
    file_exists = os.path.exists(CSV_FILE_PATH)
    with open(CSV_FILE_PATH, mode='a', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=["지원자 이름", "나이", "지원팀명", "지원파트", "직급", "경력", "핵심기술력", "특징"])
        if not file_exists:
            writer.writeheader()
        writer.writerow(result)

async def update_faiss_with_result(result):
    lock = threading.Lock()
    with lock:
        # FAISS 벡터 저장소 업데이트
        documents = [
            Document(page_content=f"지원자 이름: {result.get('지원자 이름', '')}\n나이: {result.get('나이', '')}\n경력: {result.get('경력', '')}\n핵심기술력: {result.get('핵심기술력', '')}\n특징: {result.get('특징', '')}\n지원팀명: {result.get('지원팀명', '')}\n지원파트: {result.get('지원파트', '')}\n직급: {result.get('직급', '')}", metadata={"file": result.get("지원자 이름", "")})
        ]
        await vector_store.aadd_documents(documents)
        try:
            vector_store.save_local(FAISS_INDEX_PATH)
            logging.info("Saved FAISS index to disk.")
        except Exception as e:
            logging.error(f"Failed to save FAISS index: {e}")

if __name__ == "__main__":
    event_handler = ResumeHandler()
    observer = Observer()
    observer.schedule(event_handler, path=DIRECTORY_TO_WATCH, recursive=False)
    observer.start()

    # 파일 처리 스레드 시작
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    processing_thread = threading.Thread(target=lambda: loop.run_until_complete(process_files_from_queue()), daemon=True)
    processing_thread.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        asyncio.run(file_queue.put(None))  # 큐를 종료하기 위해 None 추가
        processing_thread.join()
    observer.join()
